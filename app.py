import pyautogui
from time import sleep
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
import customtkinter as ctk
import os
import sys
import threading
import requests
from tkinter import messagebox  
from PIL import Image, ImageTk

if getattr(sys, 'frozen', False):
    application_path = sys._MEIPASS
else:
    application_path = os.path.dirname(os.path.abspath(__file__))

maringa_file = os.path.join(application_path, 'codigos_maringa.xlsx')
tapejara_file = os.path.join(application_path, 'codigos_tapejara.xlsx')

wb_debitos_maringa = openpyxl.load_workbook('codigos_maringa.xlsx')
sheet_debitos_maringa = wb_debitos_maringa['Planilha1']

wb_debitos_tapejara = openpyxl.load_workbook('codigos_tapejara.xlsx')
sheet_debitos_tapejara = wb_debitos_tapejara['Planilha1']

wb_resultado = openpyxl.Workbook()
sheet_resultado = wb_resultado.active
sheet_resultado.title = "Empresas Sem Débitos"
sheet_resultado.append(['Nome da Empresa', 'Código Municipal', 'Mensagem'])

def atualizar_ultima_linha_maringa():
    ultima_linha_maringa = ler_progresso_maringa()
    ultima_maringa.configure(text=f"(Última linha processada: {ultima_linha_maringa})")


def atualizar_ultima_linha_tapejara():
    ultima_linha_tapejara = ler_progresso_tapejara()
    ultima_tapejara.configure(text=f"(Última linha processada: {ultima_linha_tapejara})")

def ler_progresso_maringa():
    if os.path.exists("progresso_maringa.txt"):
        with open("progresso_maringa.txt", "r") as file:
            return int(file.read().strip())
    return 2  # Se não houver progresso registrado, começa da linha 2

# Função para salvar o progresso
def salvar_progresso_maringa(linha):
    with open("progresso_maringa.txt", "w") as file:
        file.write(str(linha))


def ler_progresso_tapejara():
    arquivo_progresso = "progresso_tapejara.txt"
    if os.path.exists(arquivo_progresso):
        with open(arquivo_progresso, "r") as file:
            return int(file.read().strip())
    else:
        # Cria o arquivo de progresso se não existir
        with open(arquivo_progresso, "w") as file:
            file.write("2")  # Inicia com a linha 2
        return 2

def salvar_progresso_tapejara(linha):
    with open("progresso_tapejara.txt", "w") as file:
        file.write(str(linha))

def pegar_debitos_maringa(linha_inicial_maringa):
    if threading.current_thread().name == "MainThread":
        print("Executando na thread principal")
    else:
        print("Executando em uma thread separada")  


    driver = webdriver.Chrome()
    driver.get('https://tributos.maringa.pr.gov.br/portal-contribuinte/consulta-debitos')
    
    ultima_linha_processada_maringa = ler_progresso_maringa()
    
    total_linhas = sheet_debitos_maringa.max_row
    for linha in sheet_debitos_maringa.iter_rows(min_row=linha_inicial_maringa, max_row=5):
            salvar_progresso_maringa(linha[0].row) 
            wb_resultado.save('empresas_sem_debitos_maringa.xlsx')
            nome_empresa_maringa = linha[0].value
            codigo_municipal_maringa = linha[1].value

            select_element = WebDriverWait(driver, 15).until(
                    EC.element_to_be_clickable((By.XPATH, "//select[@id='select-filter']"))
                )
            
            select = Select(select_element)
            select.select_by_value("1")  
                
            sleep(1)

            if not codigo_municipal_maringa:
                continue

            campo_cod_municipal = WebDriverWait(driver, 20).until(
                EC.visibility_of_element_located((By.XPATH, "(//input[@placeholder='Digite o cadastro...'])[2]"))
            )
            campo_cod_municipal.clear()
            sleep(1)
            campo_cod_municipal.send_keys(str(codigo_municipal_maringa))

            pyautogui.press('TAB')
            sleep(1)
            pyautogui.press('ENTER')
            
            sleep(2)
        
            try:
                empresa_sem_debitos = WebDriverWait(driver, 15).until(
                    EC.visibility_of_element_located((By.XPATH, "//article[@class='info mt-xs']"))
                )
                if empresa_sem_debitos:
                    sheet_resultado.append([nome_empresa_maringa, codigo_municipal_maringa, 'Empresa sem débitos'])
                    print(f"Empresa {nome_empresa_maringa} SEM débitos.")
            except:
                print(f"Empresa {nome_empresa_maringa} COM débitos.") 
                try: 
                    label = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "label.checkbox-item-label"))
                        )
                    label.click()
                except  Exception as e:
                     print(f"Ocorreu um erro ao tentar clicar no checkbox: {e}")
                
                folder_path = os.path.join(r'C:\Users\Logika\Desktop\Boletos_Tapejara', nome_empresa_maringa)

                if not os.path.exists(folder_path):
                    os.makedirs(folder_path)

                try:
                    boleto = WebDriverWait(driver, 3).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "em.fa.fa-file-text-o")) #botao do boleto com JS
                    )
                    driver.execute_script("arguments[0].click();", boleto)

                    # Aguardar a abertura de uma nova janela (pop-up ou aba)
                    WebDriverWait(driver, 10).until(lambda driver: len(driver.window_handles) > 1)

                    # Mudar para a nova janela (pop-up ou aba)
                    driver.switch_to.window(driver.window_handles[-1])

                    boleto_link = driver.current_url   #pega URL DO LINK

                    if boleto_link:
                        print("Link do boleto encontrado:", boleto_link)
                        
                        # Fazer o download do boleto usando requests
                        response = requests.get(boleto_link)

                        if response.status_code == 200:
                            nome_arquivo = f"{nome_empresa_maringa}_boleto.pdf"
                            caminho_arquivo = os.path.join(folder_path, nome_arquivo)

                        
                            with open(caminho_arquivo, 'wb') as f:
                                f.write(response.content)

                            print(f"Boleto salvo em: {caminho_arquivo}")
                        else:
                            print("Falha ao baixar o boleto. Status code:", response.status_code)

                    else:
                        print("Não foi possível obter o link do boleto.")

                except Exception as e:
                    print(f"Ocorreu um erro ao tentar acessar o link do boleto: {e}")
                    
                pyautogui.hotkey('ctrl', 'w')
                sleep(2)
                WebDriverWait(driver, 10).until(lambda driver: len(driver.window_handles) > 0)

                # Alternar para a última janela que permanece aberta
                driver.switch_to.window(driver.window_handles[-1])
            
    

    wb_resultado.save('empresas_sem_debitos_maringa.xlsx')
    driver.quit()  
    
    

def pegar_debitos_tapejara(linha_inicial_tapejara):
    
    driver = webdriver.Chrome()
    driver.get('https://tapejara.eloweb.net/portal-contribuinte/consulta-debitos')

    total_linhas = sheet_debitos_tapejara.max_row

    for linha in sheet_debitos_tapejara.iter_rows(min_row=linha_inicial_tapejara, max_row=5):
            print(f"Processando linha: {linha[0].row}") 
            salvar_progresso_tapejara(linha[0].row)  # Salva o progresso atual
            wb_resultado.save('empresas_sem_debitos_tapejara.xlsx')
            nome_empresa_tapejara = linha[0].value
            codigo_municipal_tapejara = linha[1].value
            
            selecionar_elemento = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//select[@id='select-filter']"))
            )
            
            selecionar = Select(selecionar_elemento)
            selecionar.select_by_value("1")  
            sleep(1) 
            
            if not codigo_municipal_tapejara:
                continue

            campo_cod_municipal = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "(//input[@placeholder='Digite o cadastro...'])[2]"))
            )
            campo_cod_municipal.clear()
            sleep(1)
            campo_cod_municipal.send_keys(str(codigo_municipal_tapejara))

            pyautogui.press('TAB')
            sleep(1)
            pyautogui.press('ENTER')
            sleep(2)
        
            empresa_sem_debitos = EC.visibility_of_element_located((By.XPATH,"//article[@class='info mt-xs']"))
            sleep(2)
    
            try:
                empresa_sem_debitos = WebDriverWait(driver, 3).until(
                    EC.visibility_of_element_located((By.XPATH, "//article[@class='info mt-xs']"))
                )
                if empresa_sem_debitos:
                    sheet_resultado.append([nome_empresa_tapejara, codigo_municipal_tapejara, 'Empresa sem débitos'])
                    print(f"Empresa {nome_empresa_tapejara} SEM débitos.")
            except:
                print(f"Empresa {nome_empresa_tapejara} COM débitos.") 
                try: 
                    label = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "label.checkbox-item-label"))
                        )
                    label.click()
                except  Exception as e:
                     print(f"Ocorreu um erro ao tentar clicar no checkbox: {e}")
                
                folder_path = os.path.join(r'C:\Users\Logika\Desktop\Boletos_Tapejara', nome_empresa_tapejara)

                if not os.path.exists(folder_path):
                    os.makedirs(folder_path)

                try:
                    boleto = WebDriverWait(driver, 3).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "em.fa.fa-file-text-o")) #botao do boleto com JS
                    )
                    driver.execute_script("arguments[0].click();", boleto)

                    # Aguardar a abertura de uma nova janela (pop-up ou aba)
                    WebDriverWait(driver, 10).until(lambda driver: len(driver.window_handles) > 1)

                    # Mudar para a nova janela (pop-up ou aba)
                    driver.switch_to.window(driver.window_handles[-1])

                    boleto_link = driver.current_url   #pega URL DO LINK

                    if boleto_link:
                        print("Link do boleto encontrado:", boleto_link)
                        
                        # Fazer o download do boleto usando requests
                        response = requests.get(boleto_link)

                        if response.status_code == 200:
                            nome_arquivo = f"{nome_empresa_tapejara}_boleto.pdf"
                            caminho_arquivo = os.path.join(folder_path, nome_arquivo)

                        
                            with open(caminho_arquivo, 'wb') as f:
                                f.write(response.content)

                            print(f"Boleto salvo em: {caminho_arquivo}")
                        else:
                            print("Falha ao baixar o boleto. Status code:", response.status_code)

                    else:
                        print("Não foi possível obter o link do boleto.")

                except Exception as e:
                    print(f"Ocorreu um erro ao tentar acessar o link do boleto: {e}")
                    
                pyautogui.hotkey('ctrl', 'w')
                sleep(2)
                WebDriverWait(driver, 10).until(lambda driver: len(driver.window_handles) > 0)


                # Alternar para a última janela que permanece aberta
                driver.switch_to.window(driver.window_handles[-1])
                

    
    driver.quit()  



#INTERFACE GRÁFICA

def verificar_thread_maringa(thread):
    if thread.is_alive():
        app.after(1000, verificar_thread_maringa, thread)
    else:
        atualizar_ultima_linha_maringa()

def verificar_thread_tapejara(thread):
    if thread.is_alive():
        app.after(1000, verificar_thread_tapejara, thread)
    else:
        atualizar_ultima_linha_tapejara()

thread_maringa = None
thread_tapejara = None

def iniciar_maringa():
    global thread_maringa
    try:
        linha_inicial_maringa = int(entry_maringa.get())
        if linha_inicial_maringa < 2:
            raise ValueError("A linha inicial deve ser maior que 1.")
        
        thread_maringa = threading.Thread(target=pegar_debitos_maringa, args=(linha_inicial_maringa,))
        thread_maringa.start()

        app.after(1000, verificar_thread_maringa, thread_maringa)

    except ValueError as e:
        messagebox.showerror("Erro", f"Entrada inválida para Maringá: {e}")


def iniciar_tapejara():
    global thread_tapejara
    try:
        linha_inicial_tapejara = int(entry_tapejara.get())
        if linha_inicial_tapejara < 2:
            raise ValueError("A linha inicial deve ser maior que 1.")
        
        thread_tapejara = threading.Thread(target=pegar_debitos_tapejara, args=(linha_inicial_tapejara,))
        thread_tapejara.start()

        app.after(1000, verificar_thread_tapejara, thread_tapejara)

    except ValueError as e:
        messagebox.showerror("Erro", f"Entrada inválida para Tapejara: {e}")
    
# Configuração do tema da interface gráfica

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")

app = ctk.CTk()
app.title("Office automation")
screen_width = app.winfo_screenwidth()
screen_height = app.winfo_screenheight()
app.geometry(f"{screen_width}x{screen_height-40}+0+0")


# Título
title_label = ctk.CTkLabel(
    app, 
    text="Controle de Débitos Municipais", 
    font=("Helvetica", 30, "bold"), 
    text_color="#ffffff"
)
title_label.pack(pady=20)


main_frame = ctk.CTkFrame(app, fg_color="transparent")
main_frame.pack(padx=20, pady=20, fill="both", expand=True)

frame_maringa = ctk.CTkFrame(main_frame)
frame_maringa.pack(side="left", padx=210, pady=30, fill="y")


img_maringa = ctk.CTkImage(Image.open("pref.maringa.png"), size=(300, 100))
label_img_maringa = ctk.CTkLabel(frame_maringa, image=img_maringa, text="")
label_img_maringa.pack(pady=10)

# Campo de entrada para Maringá
entry_maringa = ctk.CTkEntry(
    frame_maringa, 
    placeholder_text="Digite a linha inicial do Excel para Maringá", 
    font=("Helvetica", 14), 
    width=300
)
entry_maringa.pack(pady=10)

ultima_linha_processada_maringa = ler_progresso_maringa()

ultima_maringa = ctk.CTkLabel(
    frame_maringa, 
    text=f"(Última linha processada: {ultima_linha_processada_maringa})", 
    font=("Helvetica", 14), 
    width=300
)
ultima_maringa.pack(pady=10)


# Botão para Maringá
button_maringa = ctk.CTkButton(
    frame_maringa, 
    text="Iniciar prefeitura de Maringá/PR", 
    command=iniciar_maringa,
    font=("Helvetica", 14), 
    width=300, 
    height=40, 
    fg_color="#007acc",  
    hover_color="#005b99",  
)
button_maringa.pack(pady=15)

frame_tapejara = ctk.CTkFrame(main_frame)
frame_tapejara.pack(side="right", padx=145, pady=30, fill="y")

img_tapejara = ctk.CTkImage(Image.open("pref.tapejara.png"), size=(300, 100))
label_img_tapejara = ctk.CTkLabel(frame_tapejara, image=img_tapejara, text="")
label_img_tapejara.pack(pady=10)


# Campo de entrada para Tapejara
entry_tapejara = ctk.CTkEntry(
    frame_tapejara, 
    placeholder_text="Digite a linha inicial do Excel para Tapejara", 
    font=("Helvetica", 14), 
    width=300
)
entry_tapejara.pack(pady=10)

ultima_linha_processada_tapejara = ler_progresso_tapejara()

ultima_tapejara = ctk.CTkLabel(
    frame_tapejara, 
    text=f"(Última linha processada: {ultima_linha_processada_tapejara})",
    font=("Helvetica", 14), 
    width=300
)
ultima_tapejara.pack(pady=10)

# Botão para Tapejara
button_tapejara = ctk.CTkButton(
    frame_tapejara, 
    text="Iniciar prefeitura de Tapejara/PR", 
    command=iniciar_tapejara,
    font=("Helvetica", 14), 
    width=300, 
    height=40, 
    fg_color="#007acc",  
    hover_color="#005b99",  
)
button_tapejara.pack(pady=15)

# Rodapé
footer_label = ctk.CTkLabel(
    app, 
    text="Desenvolvido por Luiz Fernando Hillebrande", 
    font=("Helvetica", 10), 
    text_color="#c9c9c9"
)
footer_label.pack(side="bottom", pady=25)

# Configurações de layout da interface
app.grid_rowconfigure(0, weight=1)
app.grid_columnconfigure(0, weight=1)

# Função para sair da tela cheia
def sair_tela_cheia(event=None):
    app.attributes('-fullscreen', False)

app.bind("<Escape>", sair_tela_cheia)

# Inicia o loop da interface gráfica
app.mainloop()